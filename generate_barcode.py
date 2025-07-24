from pylibdmtx.pylibdmtx import encode
from PIL import Image as PILImage, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Side
import os
import math

def generate_datamatrix(code, output_dir='barcodes'):
    """Data Matrix 코드 이미지를 생성하는 함수"""
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Data Matrix 생성
    encoded = encode(str(code).encode('utf8'))
    
    # PIL Image로 변환
    img = PILImage.frombytes('RGB', (encoded.width, encoded.height), encoded.pixels)
    
    # 이미지 크기 조정 (1cm = 38픽셀 @ 96 DPI)
    new_size = (38, 38)
    img = img.resize(new_size, PILImage.Resampling.NEAREST)
    
    # 여백 추가
    border = 2
    
    # 바코드를 포함할 새 이미지 생성
    total_width = new_size[0] + 2*border
    total_height = new_size[1] + 2*border
    img_with_border = PILImage.new('RGB', (total_width, total_height), 'white')
    
    # 바코드 이미지 붙이기 (중앙)
    img_with_border.paste(img, (border, border))
    
    # 외곽선 그리기
    draw = ImageDraw.Draw(img_with_border)
    
    # 전체 영역 외곽선
    draw.rectangle(
        [(1, 1), 
         (total_width-2, total_height-2)],
        outline='black',
        width=1
    )
    
    # 파일 저장 (96 DPI 설정으로 1cm 크기 유지)
    filename = os.path.join(output_dir, f"datamatrix_{code}.png")
    img_with_border.save(filename, dpi=(96, 96))
    return filename

def create_sheet_with_codes(ws, start_code, end_code):
    """한 시트에 바코드 배치"""
    # 라벨 규격 (cm)
    printable_width = 10.40  # 실제 인쇄 가능 너비 (오른쪽 1cm 제외)
    
    # 바코드 크기와 간격 (cm)
    barcode_size = 1.00  # 1cm x 1cm
    horizontal_spacing = 0.8  # 바코드 간 가로 간격
    
    # 한 행에 들어갈 수 있는 바코드 수 계산
    barcodes_per_row = math.floor(printable_width / (barcode_size + horizontal_spacing))
    
    # 엑셀 셀 크기 조정
    for col in range(barcodes_per_row):
        col_letter = chr(ord('A') + col)
        ws.column_dimensions[col_letter].width = 5  # 약 1.3cm (간격 포함)
    
    # 바코드 생성 및 배치
    current_code = start_code
    row_idx = 1
    
    while current_code <= end_code:
        for col in range(barcodes_per_row):
            if current_code > end_code:
                break
                
            # 바코드 생성
            image_path = generate_datamatrix(current_code)
            
            # 이미지 삽입
            img = Image(image_path)
            img.width = 38  # 1cm
            img.height = 38  # 1cm (텍스트 제거로 높이 감소)
            
            # 엑셀에 이미지 추가
            cell = f"{chr(ord('A') + col)}{row_idx}"
            ws.add_image(img, cell)
            
            current_code += 1
        
        # 행 높이 설정
        ws.row_dimensions[row_idx].height = 38  # 1cm
        row_idx += 1

def create_label_sheets(start_code, num_codes, codes_per_sheet=100, output_file='barcodes.xlsx'):
    """여러 시트로 구성된 엑셀 파일 생성"""
    wb = Workbook()
    
    # 첫 번째 시트는 기본 생성되므로 삭제
    wb.remove(wb.active)
    
    # 시트별로 바코드 생성
    for sheet_start in range(start_code, start_code + num_codes, codes_per_sheet):
        sheet_end = min(sheet_start + codes_per_sheet - 1, start_code + num_codes - 1)
        
        # 시트 이름 설정 (예: "6000-6099")
        sheet_name = f"{sheet_start}-{sheet_end}"
        ws = wb.create_sheet(sheet_name)
        
        # 시트에 바코드 추가
        create_sheet_with_codes(ws, sheet_start, sheet_end)
    
    # 엑셀 파일 저장
    try:
        wb.save(output_file)
        print(f"엑셀 파일이 생성되었습니다: {output_file}")
        print(f"총 {math.ceil(num_codes/codes_per_sheet)}개의 시트가 생성되었습니다.")
    except PermissionError:
        print("엑셀 파일이 다른 프로그램에서 열려있습니다. 파일을 닫고 다시 시도해주세요.")
        return

def main():
    # 6000번부터 7999번까지 2000개의 바코드를 100개씩 시트로 분리
    create_label_sheets(0, 1000, codes_per_sheet=1000)

if __name__ == '__main__':
    main() 

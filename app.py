from pylibdmtx.pylibdmtx import encode
from PIL import Image as PILImage, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Side
import os
import math

def generate_datamatrix(code, output_dir='barcodes', add_t=False):
    """Data Matrix 코드 이미지를 생성하는 함수"""
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Data Matrix 생성
    encoded = encode(str(code).encode('utf8'))
    
    # PIL Image로 변환
    img = PILImage.frombytes('RGB', (encoded.width, encoded.height), encoded.pixels)
    
    # 이미지 크기 조정 (1.2cm = 45픽셀 @ 96 DPI)
    new_size = (45, 45)
    img = img.resize(new_size, PILImage.Resampling.NEAREST)
    
    # 여백 추가
    border = 2
    
    # 바코드를 포함할 새 이미지 생성
    total_width = new_size[0] + 2*border
    total_height = new_size[1] + 2*border
    img_with_border = PILImage.new('RGB', (total_width, total_height), 'white')
    
    # 바코드 이미지 붙이기 (중앙)
    img_with_border.paste(img, (border, border))
    
    # 외곽선 그리기
    draw = ImageDraw.Draw(img_with_border)
    
    # 전체 영역 외곽선
    draw.rectangle(
        [(1, 1), 
         (total_width-2, total_height-2)],
        outline='black',
        width=1
    )

    # 'T' 글자 추가
    if add_t:
        try:
            font = ImageFont.truetype("arial.ttf", 12)
        except IOError:
            font = ImageFont.load_default()
        draw.text((total_width - 10, 2), "T", fill="black", font=font)

    # 파일 저장 (96 DPI 설정으로 1.2cm 크기 유지)
    filename_suffix = "_t" if add_t else ""
    filename = os.path.join(output_dir, f"datamatrix_{code}{filename_suffix}.png")
    img_with_border.save(filename, dpi=(96, 96))
    return filename

def create_sheet_with_codes(ws, start_code, end_code):
    """한 시트에 바코드 배치"""
    # 라벨 규격 (cm)
    printable_width = 20.0  # A4 너비 기준, 여유있게 설정
    
    # 바코드 크기와 간격 (cm)
    barcode_size = 1.2  # 1.2cm x 1.2cm
    horizontal_spacing = 0.5  # 바코드 간 가로 간격
    
    # 한 행에 들어갈 원본 바코드 수 계산 (원본 + T-복사본 세트)
    barcodes_per_row_single = math.floor(printable_width / (barcode_size * 2 + horizontal_spacing * 2))
    total_columns = barcodes_per_row_single * 2

    # 엑셀 셀 크기 조정
    for col in range(total_columns):
        col_letter = chr(ord('A') + col)
        ws.column_dimensions[col_letter].width = 6.5

    # 바코드 생성 및 배치
    current_code = start_code
    row_idx = 1
    
    while current_code <= end_code:
        # 한 행에 바코드 배치
        for i in range(barcodes_per_row_single):
            if current_code > end_code:
                break
            
            # 원본 바코드 생성
            image_path_orig = generate_datamatrix(current_code, add_t=False)
            img_orig = Image(image_path_orig)
            img_orig.width = 45
            img_orig.height = 45
            
            # T-복사본 바코드 생성
            image_path_t = generate_datamatrix(current_code, add_t=True)
            img_t = Image(image_path_t)
            img_t.width = 45
            img_t.height = 45

            # 엑셀에 이미지 추가
            # 원본 (왼쪽)
            cell_orig = f"{chr(ord('A') + i)}{row_idx}"
            ws.add_image(img_orig, cell_orig)
            
            # T-복사본 (오른쪽 대칭 위치)
            cell_t = f"{chr(ord('A') + total_columns - 1 - i)}{row_idx}"
            ws.add_image(img_t, cell_t)

            current_code += 1

        # 행 높이 설정
        ws.row_dimensions[row_idx].height = 48  # 1.2cm 보다 약간 크게
        row_idx += 1

def create_label_sheets(start_code, num_codes, codes_per_sheet=100, output_file='barcodes.xlsx'):
    """여러 시트로 구성된 엑셀 파일 생성"""
    wb = Workbook()
    
    # 첫 번째 시트는 기본 생성되므로 삭제
    wb.remove(wb.active)
    
    # 시트별로 바코드 생성
    for sheet_start in range(start_code, start_code + num_codes, codes_per_sheet):
        sheet_end = min(sheet_start + codes_per_sheet - 1, start_code + num_codes - 1)
        
        # 시트 이름 설정 (예: "6000-6099")
        sheet_name = f"{sheet_start}-{sheet_end}"
        ws = wb.create_sheet(sheet_name)
        
        # 시트에 바코드 추가
        create_sheet_with_codes(ws, sheet_start, sheet_end)
    
    # 엑셀 파일 저장
    try:
        wb.save(output_file)
        print(f"엑셀 파일이 생성되었습니다: {output_file}")
        print(f"총 {math.ceil(num_codes/codes_per_sheet)}개의 시트가 생성되었습니다.")
    except PermissionError:
        print("엑셀 파일이 다른 프로그램에서 열려있습니다. 파일을 닫고 다시 시도해주세요.")
        return

def main():
    # 6000번부터 7999번까지 2000개의 바코드를 100개씩 시트로 분리
    create_label_sheets(0, 1000, codes_per_sheet=1000)

if __name__ == '__main__':
    main() 
